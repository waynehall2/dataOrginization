import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill

# Load the Excel file
file_name = 'inputFileName.xlsx'
df = pd.read_excel(file_name)

# Delete specified columns
columns_to_delete = ['Start time', 'Email', 'Name', 'Last modified time']
df = df.drop(columns=columns_to_delete)

# Insert new columns at specific positions
new_columns = {
    4: 'Cloud',
    12: 'Cybersecurity',
    30: 'PM',
    39: 'Containerization',
    46: 'Programming',
    59: 'Data Analytics',
    65: 'Reverse Engineering',
    81: 'Operating Systems',
    93: 'Testing',
    101: 'Zero Trust', 
    107: 'Technical Writing',
    109: 'Biometrics',
    111: 'CNO Development',
    113: 'Certifications',
    129: 'Additional Certification'
}
for position, column_name in new_columns.items():
    df.insert(position, column_name, '')

# Save the modified DataFrame to a new Excel file
output_file_name = 'outPutFileName.xlsx'
df.to_excel(output_file_name, index=False)

# Load the workbook using openpyxl
wb = load_workbook(output_file_name)

# Rename the first sheet to "Skills"
wb.active.title = "Skills"

# Create new sheets and set their positions after the first sheet
sheet_names = [
    "Cloud",
    "Cybersecurity",
    "PM",
    "Containerization",
    "Programming",
    "Data Analytics",
    "Reverse Engineering",
    "Operating Systems",
    "Testing",
    "Zero Trust",
    "Technical Writing",
    "Biometrics",
    "CNO Development", 
    "Certifications",
    "Additional Certification"
]
for index, name in enumerate(sheet_names, start=2):
    sheet = wb.create_sheet(name)
    wb.move_sheet(sheet, index)

# Function to copy data from source to target sheet
def copy_data(source_sheet, target_sheet, columns):
    for row_index, column_data in enumerate(zip(*columns), start=1):
        for col_index, cell in enumerate(column_data, start=1):
            cell_value = cell.value
            target_sheet.cell(row=row_index, column=col_index, value=cell_value)

# Dictionary mapping sheet names to columns
sheet_columns = {
    "Cloud": [1, 2, 3, 4, 6, 7, 8, 9, 10, 11, 12],
    "Cybersecurity": [1, 2, 3, 4] + list(range(14, 31)),
    "PM": [1, 2, 3, 4] + list(range(32, 40)),
    "Containerization": [1, 2, 3, 4] + list(range(41, 47)),
    "Programming": [1, 2, 3, 4] + list(range(48, 60)),
    "Data Analytics": [1, 2, 3, 4] + list(range(61, 66)),
    "Reverse Engineering": [1, 2, 3, 4] + list(range(67, 82)),
    "Operating Systems": [1, 2, 3, 4] + list(range(83, 94)),
    "Testing": [1, 2, 3, 4] + list(range(95, 102)),
    "Zero Trust": [1, 2, 3, 4] + list(range(103, 108)),
    "Technical Writing": [1, 2, 3, 4, 109],
    "Biometrics": [1, 2, 3, 4, 111],
    "CNO Development": [1, 2, 3, 4, 113],
    "Certifications": [1, 2, 3, 4] + list(range(115, 130)),
    "Additional Certification": [1, 2, 3, 4, 131]
}

# Copy data from "Skills" to each sheet
skills_sheet = wb['Skills']
for sheet_name, columns in sheet_columns.items():
    target_sheet = wb[sheet_name]
    selected_columns = [skills_sheet[f'{get_column_letter(col)}'] for col in columns]
    copy_data(skills_sheet, target_sheet, selected_columns)


# Apply formulas to the "Skills" sheet only
sheet = wb['Skills']

# Define a slightly darker blue fill pattern using a hexadecimal color code
slightly_darker_blue_fill = PatternFill(start_color="6699CC", end_color="6699CC", fill_type="solid")

# Apply the slightly darker blue fill pattern to all cells in row 1
for cell in sheet[1]:
    cell.fill = slightly_darker_blue_fill

# Determine the last column with data in row 1
last_column_with_data = sheet.max_column

# Add a filter to row 1 for all columns with data
sheet.auto_filter.ref = f"A1:{get_column_letter(last_column_with_data)}{1}"

# Freeze the pane at row 1 and column 4
sheet.freeze_panes = "E2"

# Define the ranges where you want to apply the formulas
formulas = {
    'E': '=IF(COUNTIF(F{row}:L{row},"*")>0, "X","")',
    'M': '=IF(COUNTIF(N{row}:AD{row},"*")>0, "X","")',
    'AE': '=IF(COUNTIF(AF{row}:AM{row},"*")>0, "X","")',
    'AN': '=IF(COUNTIF(AO{row}:AT{row},"*")>0, "X","")',
    'AU': '=IF(COUNTIF(AV{row}:BG{row},"*")>0, "X","")',
    'BH': '=IF(COUNTIF(BI{row}:BM{row},"*")>0, "X","")',
    'BN': '=IF(COUNTIF(BO{row}:CC{row},"*")>0, "X","")',
    'CD': '=IF(COUNTIF(CE{row}:CO{row},"*")>0, "X","")',
    'CP': '=IF(COUNTIF(CQ{row}:CW{row},"*")>0, "X","")',
    'CX': '=IF(COUNTIF(CY{row}:DC{row},"*")>0, "X","")',
    'DD': '=IF(COUNTIF(DE{row}:DE{row},"*")>0, "X","")',
    'DF': '=IF(COUNTIF(DG{row}:DG{row},"*")>0, "X","")',
    'DH': '=IF(COUNTIF(DI{row}:DI{row},"*")>0, "X","")',
    'DJ': '=IF(COUNTIF(DK{row}:DY{row},"*")>0, "X","")',
    'DZ': '=IF(COUNTIF(EA{row}:EA{row},"*")>0, "X","")',
}

# Apply the formulas to each row individually, but only if the row has data
for column, formula in formulas.items():
    for row in range(2, sheet.max_row + 1): # Adjusted to dynamically find the last row with data
        cell = f'{column}{row}'
        # Check if the cell in the first column of the current row is not empty
        if sheet[f'A{row}'].value is not None and str(sheet[f'A{row}'].value).strip() != '':
            sheet[cell] = formula.format(row=row)

# Define a very light shade of blue fill pattern using a hexadecimal color code
very_light_blue_fill = PatternFill(start_color="F0F8FF", end_color="F0F8FF", fill_type="solid")

# Center all data in the cells
center_alignment = Alignment(horizontal='center', vertical='center')

# Iterate over each sheet in the workbook
for sheet_name in wb.sheetnames:
    sheet_obj = wb[sheet_name]
    
    # Delete rows that do not contain data in any column after column 4
    for row in range(sheet_obj.max_row, 0, -1): # Iterate backwards to avoid index shifting
        if all(cell.value is None or str(cell.value).strip() == '' for cell in sheet_obj[row][4:]):
            sheet_obj.delete_rows(row)

    
    # Apply the slightly darker blue fill pattern to all cells in row 1
    for cell in sheet_obj[1]:
        cell.fill = slightly_darker_blue_fill
    
    # Determine the last column with data in row 1
    last_column_with_data = sheet_obj.max_column
    
    # Add a filter to row 1 for all columns with data
    sheet_obj.auto_filter.ref = f"A1:{get_column_letter(last_column_with_data)}{1}"
    
    # Freeze the pane at row 1 and column 4
    sheet_obj.freeze_panes = "E2"
    
    # Adjust column width to fit the content
    for column in sheet_obj.columns:
        max_length = max(len(str(cell.value)) for cell in column if cell.value is not None)
        adjusted_width = (max_length + 5)
        sheet_obj.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width
        
    # Apply the very light shade of blue fill pattern to every other row starting from row 3, moving downwards
    for row_index in range(3, sheet_obj.max_row + 1, 2): # Start from row 3 and move downwards in steps of 2
        for cell in sheet_obj[row_index]:
            cell.fill = very_light_blue_fill
    
    # Center all data in the cells
    for row in sheet_obj.iter_rows():
        for cell in row:
            cell.alignment = center_alignment

# Save the workbook
wb.save(output_file_name)
